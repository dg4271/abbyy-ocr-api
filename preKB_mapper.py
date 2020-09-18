import json
import requests
import pandas as pd
import csv
from collections import defaultdict

from table_semantic_tagging.semantic_tagging import * 
from table_semantic_tagging.value_regularizer import * 
from xml_to_table.table_extractor import * 

class GroupInfo:
    def __init__(self, groupName="", timeline="", attribute="", value_1="", unit_1="", 
                 value_2="", unit_2="", std="", min="", max="", terms=""):
        self.groupName = groupName
        self.timeline = timeline
        self.attribute = attribute
        #self.mean = mean
        self.value_1 = value_1
        self.unit_1 = unit_1
        self.value_2 = value_2
        self.unit_2 = unit_2
        self.std = std
        self.min = min
        self.max = max
        self.terms = terms

    def toJson(self):
        return json.dumps(self, default=lambda o: o.__dict__)

class PValue:
    def __init__(self, isSignificant="", value="", originalCell=""):
        self.isSignificant = isSignificant
        self.value = value
        self.originalCell = originalCell


class GeneInfo:
    def __init__(self, groupName="", gene="", snp="", bm="", bmc=""):
        self.groupName = groupName
        self.gene = gene
        self.snp = snp
        self.bm = bm
        self.bmc = bmc

    def toJson(self):
        return json.dumps(self, default=lambda o: o.__dict__)


class TreatmentInfo:
    def __init__(self, groupName="", treatment="", treatmentType="", idu="", \
        ido="", pa="", dic="", di="", dc="", bm="", bmc=""):
        self.groupName = groupName
        self.treatment = treatment
        self.treatmentType = treatmentType
        self.idu = idu
        self.ido = ido
        self.pa = pa
        self.dic = dic
        self.di = di
        self.dc = dc
        self.bm = bm
        self.bmc = bmc

    def toJson(self):
        return json.dumps(self, default=lambda o: o.__dict__)


class TreatmentGeneInfo:
    def __init__(self, groupName="", treatment="", treatmentType="", gene="", \
        snp="", idu="", ido="", pa="", dic="", di="", dc="", bm="", bmc=""):
        self.groupName = groupName
        self.treatment = treatment
        self.treatmentType = treatmentType
        self.gene = gene
        self.snp = snp
        self.idu = idu
        self.ido = ido
        self.pa = pa
        self.dic = dic
        self.di = di
        self.dc = dc
        self.bm = bm
        self.bmc = bmc

    def toJson(self):
        return json.dumps(self, default=lambda o: o.__dict__)


class ExperimentInfo:
    def __init__(self, controlGroup="", controlTimeline="", experimentGroup="", \
        experimentTimeline="", attribute="", pvalue="", isSignificant="", isDecrease=""):
        self.controlGroup = controlGroup
        self.controlTimeline = controlTimeline
        self.experimentGroup = experimentGroup
        self.experimentTimeline = experimentTimeline
        self.attribute = attribute
        self.pvalue = pvalue
        self.isSignificant = isSignificant
        self.isDecrease = isDecrease

    def toJson(self):
        return json.dumps(self, default=lambda o: o.__dict__)


class Data:
    def __init__(self, sourceType="", groupInfo=[], treatmentGeneInfo=[], experimentInfo=[]):
        self.sourceType = sourceType
        self.groupInfo = groupInfo
        self.treatmentGeneInfo = treatmentGeneInfo
        self.experimentInfo = experimentInfo

    def toJson(self):
        return json.dumps(self, default=lambda o: o.__dict__)
    

class PreKB:
    resultCode = 0
    data = []

    def __init__(self, resultCode=0, data=[]):
        self.resultCode = resultCode
        self.data = data

    def toJson(self):
        return json.dumps(self, default=lambda o: o.__dict__, indent=3)


def join_treatment_gene(treatmentInfos, geneInfos):
    
    """
    비정형의 output인 treatmentInfo와 geneInfo를 join하여 treatmentGeneInfo 생성
    TreatmentInfo -> dictionary -> DataFrame -> outer join

    1. 둘 다 비어있는 경우 => []
    2. geneInfos만 비어있을 경우 => treatmentInfos 정보를 기반으로 treamentGeneInfo 생성
    3. treatmentInfos만 비어있을 경우 => geneInfos 정보를 기반으로 treamentGeneInfo 생성
    4. treatmentInfos와 geneInfos의 길이가 모두 0이 아닌 경우 => outer join
    """

    print(len(treatmentInfos))
    print(len(geneInfos))

    treatmentGeneInfo = []

    if len(treatmentInfos)==0 and len(geneInfos)==0:
        return []
    elif len(treatmentInfos)==0 and len(geneInfos) != 0:

        for _geneInfo in geneInfos:
        
            if (_geneInfo.gene or _geneInfo.snp) and _geneInfo.bm:
                treatmentGeneInfo.append(TreatmentGeneInfo(_geneInfo.groupName, "", \
                    "", _geneInfo.gene, _geneInfo.snp, "", \
                    "",  "", "", "", \
                    "", _geneInfo.bm, _geneInfo.bmc))
            

        return treatmentGeneInfo
    elif len(treatmentInfos)!=0 and len(geneInfos) == 0:

        for _treatmentInfo in treatmentInfos:
        
            if _treatmentInfo.treatment and (_treatmentInfo.pa or _treatmentInfo.bm):
                treatmentGeneInfo.append(TreatmentGeneInfo(_treatmentInfo.groupName, _treatmentInfo.treatment, \
                    _treatmentInfo.treatmentType, "", "", _treatmentInfo.idu, \
                    _treatmentInfo.ido,  _treatmentInfo.pa, _treatmentInfo.dic, _treatmentInfo.di, \
                    _treatmentInfo.dc, _treatmentInfo.bm, _treatmentInfo.bmc))
        
        return treatmentGeneInfo
    else:
        treatment_dict = defaultdict(list)
        gene_dict = defaultdict(list)

        for _treatmentInfo in treatmentInfos:
            for k, v in _treatmentInfo.__dict__.items():
                treatment_dict[k].append(v)

        for _geneInfo in geneInfos:
            for k, v in _geneInfo.__dict__.items():
                gene_dict[k].append(v)

        treament_df = pd.DataFrame.from_dict(treatment_dict, orient='index').T
        gene_df = pd.DataFrame.from_dict(gene_dict, orient='index').T

        

        df_merge_outer = pd.merge(treament_df, gene_df, how='outer', on=['groupName', 'bm', 'bmc'])
        df_merge_outer.fillna("", inplace = True)
        
        ## for debugging
        # print(treament_df.head())
        # print(gene_df.head())
        # print(df_merge_outer.head())

        for _, row in df_merge_outer.iterrows():

            is_gene_or_snp = False
            is_pa_or_bm = False
            is_treat = False

            if row["gene"] or row["snp"]:
                is_gene_or_snp = True
            
            if row["pa"] or row["bm"]:
                is_pa_or_bm = True
            
            if row["treatment"]:
                is_treat = True

            
            if (is_gene_or_snp and is_pa_or_bm) or (is_gene_or_snp and is_treat) or (is_pa_or_bm and is_treat):
                treatmentGeneInfo.append(TreatmentGeneInfo(row["groupName"], row["treatment"], \
                                        row["treatmentType"], row["gene"], row["snp"], row["idu"], row["ido"], \
                                        row["pa"], row["dic"], row["di"], row["dc"], row["bm"], row["bmc"]))
            else:
                print("Filterd row from text ")
                print(row)
                continue

        return treatmentGeneInfo
  
def req_sents(doc_id):
    SENT_API = "http://211.109.9.77:10600/cjpaper/sent"

    data = {
        "id":doc_id,
        "index": 0,
        "count": 10000
    }

    try:
        r = requests.post(SENT_API, json=data)
        return r.json()
    except:
        return None

def req_methods(json_data):
    BIOBERT_API = "http://211.109.9.141:3164/biobert"
    try:
        r = requests.post(url=BIOBERT_API, json=json_data)
        return r.json()
    except:
        return None

def from_unstructured(doc_id):

    data = []
    
    sents_json = req_sents(doc_id)
    if sents_json is None:
        print("result of sents is None ", doc_id)
        return data
    
    #print(sents_json)

    biobert_result = req_methods(sents_json)
    if biobert_result is None:
        print("result of biobert is None ", doc_id)
        return data

    
    BASE_PATH = "/data1/saltlux/CJPoc/table-extraction/final_demo/task1_30/task1_biobert_result/"
    with open(BASE_PATH+"biobert_"+doc_id+".json", 'w') as json_file:
        json.dump(biobert_result, json_file, indent=4)

    groupInfo = []
    geneInfo = []
    treatmentInfo = []
    treatmentGeneInfo = []
    experimentInfo = []

    # try: 
    # geneInfo
    try:
        for _geneInfo in biobert_result["geneInfo"]:
            #print(_geneInfo)
            groupName = _geneInfo["groupName"]
            genes = _geneInfo["genes"]
            snp = _geneInfo["snp"]
            bmInfos = _geneInfo["bmInfo"]

            """
            genes와 bmInfos의 리스트의 원소 존재 유무에 따라 4가지 경우,,
            """
            if len(bmInfos) != 0 and len(genes) != 0:
                for bmInfo in _geneInfo["bmInfo"]:
                    for gene in genes:
                        geneInfo.append(GeneInfo(groupName, gene, snp, bmInfo["bm"], bmInfo["bmc"]))
            elif len(bmInfos) == 0 and len(genes) != 0:
                for gene in genes:
                    geneInfo.append(GeneInfo(groupName, gene, snp, "", ""))
            elif len(bmInfos) != 0 and len(genes) == 0:
                for bmInfo in _geneInfo["bmInfo"]:
                    geneInfo.append(GeneInfo(groupName, "", snp, bmInfo["bm"], bmInfo["bmc"]))    
            else:
                geneInfo.append(GeneInfo(groupName, "", snp, "", ""))    
    except Exception as e:
        print("geneInfo parsing error")
        print(biobert_result["geneInfo"])  
        print("Error: ", e)                  
        

    # groupInfo
    try:
        for _groupInfo in biobert_result["groupInfo"]:

            groupName = _groupInfo["groupName"]
            timeline = _groupInfo["timeLine"]
            attribute = _groupInfo["attribute"]
            mean = _groupInfo["mean"]
            std = _groupInfo["std"]
            unit = _groupInfo["unit"]
            terms = "|".join(_groupInfo["terms"])

            groupInfo.append(GroupInfo(groupName, timeline, attribute, mean, "", "", "", std, "", "", terms))
    except Exception as e:
        print("groupInfo parsing error")
        print(biobert_result["groupInfo"])
        print("Error: ", e)


    # treatementInfo
    try:
        for _treatmentInfo in biobert_result["treatmentInfo"]:
            #print(_treatmentInfo)
            groupName = _treatmentInfo["groupName"]
            treatment = _treatmentInfo["treatName"]
            treatmentType = _treatmentInfo["treatType"]
            idu = "|".join(_treatmentInfo["idu"])
            ido = "|".join(_treatmentInfo["ido"])
            pa = "|".join(_treatmentInfo["pa"])
            dic = "|".join(_treatmentInfo["dic"])
            diseaseInfos = _treatmentInfo["diInfo"]
            bmInfos = _treatmentInfo["bmInfo"]

            if len(bmInfos) != 0 and len(diseaseInfos) != 0:
                for _bmInfo in bmInfos:
                    for _diseaseInfo in diseaseInfos:
                        treatmentInfo.append(TreatmentInfo(groupName, treatment, \
                            treatmentType, idu, ido, pa, dic, \
                            _diseaseInfo["di"], "|".join(_diseaseInfo["dc"]), _bmInfo["bm"], _bmInfo["bmc"]))
            elif len(bmInfos) == 0 and len(diseaseInfos) != 0:
                for _diseaseInfo in diseaseInfos:
                        treatmentInfo.append(TreatmentInfo(groupName, treatment, \
                            treatmentType, idu, ido, pa, dic, \
                            _diseaseInfo["di"], "|".join(_diseaseInfo["dc"]), "", ""))
            elif len(bmInfos) != 0 and len(diseaseInfos) == 0:
                for _bmInfo in bmInfos:
                        treatmentInfo.append(TreatmentInfo(groupName, treatment, \
                            treatmentType, idu, ido, pa, dic, \
                            "", "", _bmInfo["bm"], _bmInfo["bmc"]))
            else:
                treatmentInfo.append(TreatmentInfo(groupName, treatment, \
                            treatmentType, idu, ido, pa, dic, \
                            "", "", "", ""))
    except Exception as e:
        print("treatmentInfo parsing error")
        print(biobert_result["treatmentInfo"])
        print("Error: ", e)
    
 
    try:
        treatmentGeneInfo = join_treatment_gene(treatmentInfo, geneInfo)
    except Exception as e:
        print("join_treatment_gene error")
        print(treatmentInfo)
        print(geneInfo)
        print("Error: ", e)

    


    data.append(Data("Text", groupInfo, treatmentGeneInfo, experimentInfo))
    
    return data

def from_structured(xml_path): 
    # 사전을 불러오기 때문에 클래스 객체를 먼저 생성한다.
    st_module = Semantic()

    # Value 정제 클래스
    vr = ValueRegularizer()


    # 표 별 지식 추출 결과  
    data = []

    """
    1. ABBYY XML 결과 -> Table 정보 추출    
    """
    try:
        extractedTables = xml_to_table(xml_path)["tableViewInfos"]
    except Exception as e:
        print("xml to table error")
        raise Exception(e)

    for extractedTable in extractedTables:
        
        """
        2. Table type
          - Table type 활용하여 experimentInfo 추출 여부 판단?
        # caption = []
        # is_table = is_collect_table(table['table']) 
        # table['is_table'] = is_table
        # caption.extend(table['upcaption'])
        # caption.extend(table['downcaption'])
        # captions, row_header_list, col_header_list = get_header_value_info_new(table, caption)
        # table_type = get_type(captions, row_header_list, col_header_list)
        # table['table_type'] = table_type
        # print(extractedTable['table_type'])
        """

        # cell: <row header, col header, value> 단위로 groupInfo 추출
        # row: [ <row header, col header, value> ] 단위로 experimentInfo 추출 
        groupInfo = [] 
        experimentInfo = []
        
        for rows in extractedTable["hv"]:
            
            row_groupInfos = []
            row_p_values = []
            for cell in rows["row"]: 
                """
                3. Semantic Tagging
                  - tagged_list_header
                    : header list을 Semantic Tagging의 Input으로 받을 수 있음
                """                    
                
                if not cell["row_header"] or not cell["col_header"]:
                    print("row or col header is empty")
                    #print("cell: ", cell)
                    continue

                try: 
                    semantic_row_header = st_module.tagged_list_header(cell["row_header"])
                    semantic_col_header = st_module.tagged_list_header(cell["col_header"])
                except Exception as e:
                    print("semantic module failed")
                    #print("cell: ", cell)
                    print("Error: ", e)
                    continue
                
                """
                3. Value regularizer
                    - mean, SD, terms 추출
                """
                # regularzied_values = vr.regularize_value(cell["value"])
                try:
                    regularzied_values = vr.regularize_value_with_header(cell)
                except Exception as e:
                    print("value regularizer module error")
                    #print(cell)
                    print("Error: ", e)
                    continue

                #mean = None
                value_1 = None
                unit_1 = None
                value_2 = None
                unit_2 = None
                SD = None
                range_min = None
                range_max = None
                terms = ""
                is_p_value = False
                p_value = None 
                try:
                    # 한 cell의 value는 최대 2개까지 있는 것을 고려해 파싱
                    for regularzied_value in regularzied_values: 
                        #print(regularzied_value)
                        if regularzied_value['unit'] == 'p_value': 
                            is_p_value = True
                            p_value = PValue(regularzied_value['significant'], regularzied_value['value'], regularzied_value['originalCell'])
                            #print(p_value)

                        elif regularzied_value['type'] == "mean" or regularzied_value['type'] == "default" \
                              or regularzied_value['type'] == "count" or regularzied_value['unit'] == "count":
                            
                            #print(value_1)
                            if not value_1:
                                value_1 = regularzied_value['value']
                                unit_1 = regularzied_value['unit']
                            elif regularzied_value['unit'] == "%":
                                value_2 = regularzied_value['value']
                                unit_2 = regularzied_value['unit']

                        # elif regularzied_value['type'] == "mean" or regularzied_value['type'] == "default":
                        #     if type(regularzied_value['value']) is tuple:
                        #         mean = regularzied_value['value'][0]
                        #     else: 
                        #         mean = regularzied_value['value'] 
                            
                        elif regularzied_value['type'] == 'SD': 
                            if type(regularzied_value['value']) is tuple:
                                SD = regularzied_value['value'][0]
                            else:
                                SD = regularzied_value['value']
                        
                        elif regularzied_value['type'] == "range":
                            if type(regularzied_value['value']) is tuple:
                                range_min = regularzied_value['value'][0]
                                range_max = regularzied_value['value'][1]
                            else:
                                range_min = regularzied_value['value']

                        
                        
                except Exception as e:
                    print("regularizer parsing error", cell["value"])
                    print("Error: ", e)
                    if cell["value"]:
                        terms = cell["value"]

                """
                4. semantic type 활용하여 group / attribute 구분
                """
                if semantic_row_header and semantic_col_header:
                    if semantic_row_header["type"] == "group" or semantic_col_header["type"] == "attribute":
                        group_header, attribute_header = semantic_row_header, semantic_col_header
                    elif semantic_col_header["type"] == "group" or semantic_row_header["type"] == "attribute":
                        group_header, attribute_header = semantic_col_header, semantic_row_header
                    else:
                        group_header, attribute_header = semantic_col_header, semantic_row_header
                else:
                    print("One of semantic header is empty")
                    print("row header: ", semantic_row_header)
                    print("col header: ", semantic_col_header)
                    continue

                """
                # 5. unit 판단
                #   - attribute header의 unit을 먼저 확인
                #   - group header에만 unit이 있다면 사용
                #   - unit list중 마지막 element 사용 (header list의 마지막이 value와 연관성 높다고 판단) 
                """
                #unit = ""
                if len(attribute_header["unit"]) != 0:
                    unit_1 = attribute_header["unit"][-1]
                elif len(group_header["unit"]) != 0:
                    unit_1 = group_header["unit"][-1]

                if is_p_value and p_value:
                    row_p_values.append(p_value)
                else: 
                    row_groupInfos.append(GroupInfo(group_header["text"], group_header["timeLine"], 
                                                attribute_header["text"], value_1, unit_1, 
                                                value_2, unit_2, SD, range_min, range_max,
                                                terms))
                    
            groupInfo.extend(row_groupInfos)  

            """
            6. experiemntInfo 추출
              - 같은 행의 groupInfo 들의 관계 규칙을 파악하여 추출
              - <control group, expreimental group, p-value> 트리플 단위로 추출이 목표 
  
            """ 
  
            # Rule 1. groupInfo 2개, p-value 1개 트리플의 짝이 맞는 경우 
            if len(row_p_values) != 0 and (len(row_groupInfos) == len(row_p_values) * 2): 
                try:     
                    for i in range(len(row_p_values)):
                        control = row_groupInfos[2*i]
                        experiment = row_groupInfos[2*i+1] 
                        p_value = row_p_values[i]
                        value_difference = experiment.value_1 - control.value_1
                        isDecrease = value_difference < 0
                        
                        experimentInfo.append(ExperimentInfo(control.groupName, control.timeline, experiment.groupName, experiment.timeline,
                                                            control.attribute, p_value.value, p_value.isSignificant, isDecrease)) 
                 
                except Exception as e:
                    print("get experimentInfo error")    
                    print("Error: ", e)
                    
        # Data of a Table
        data.append(Data("Table", groupInfo, [], experimentInfo))

    return data        

def preKB_to_xlsx(preKBJson, result_path):
    
    """
    preKB json -> excel
    sheet 1: GroupInfo
    sheet 2: TreatmentGeneInfo
    sheet 3: ExperimentInfo

    """

    from openpyxl import Workbook

    write_wb = Workbook()
    groupInfo_sheet = write_wb.active
    groupInfo_sheet.title = "GroupInfo"
    treatmentGeneInfo_sheet = write_wb.create_sheet("treatmentGeneInfo")
    experimentInfo_sheet = write_wb.create_sheet("experimentInfo")
    
    # excel header
    groupInfo_sheet.append(list(GroupInfo().__dict__.keys()))
    treatmentGeneInfo_sheet.append(list(TreatmentGeneInfo().__dict__.keys()))
    experimentInfo_sheet.append(list(ExperimentInfo().__dict__.keys()))

    json_data = json.loads(preKBJson)
    data_list = json_data["data"]

    try: 
        for data in data_list:
            for groupInfo in data["groupInfo"]:
                groupInfo_sheet.append(list(groupInfo.values()))
            for treatmentGeneInfo in data["treatmentGeneInfo"]:
                treatmentGeneInfo_sheet.append(list(treatmentGeneInfo.values()))
            for experimentInfo in data["experimentInfo"]:
                experimentInfo_sheet.append(list(experimentInfo.values()))
    except ValueError as err:
        print(err)

    write_wb.save(result_path)
        
def extract_preKB(xml_path, doc_id):
    data = []
    try:
        data.extend(from_structured(xml_path))
    except Exception as e:
        print("from_structed error")
        print(e)
        data.append(Data("Table", [], [], []))
    data.extend(from_unstructured(doc_id))

    return PreKB(0, data)

def get_fileName_docID_dict(dic_file_path="/data1/saltlux/CJPoc/table-extraction/final_demo/document_id.csv"):
    fileName_docID_dict = {}
    with open(dic_file_path) as csvFile:
        reader = csv.reader(csvFile, delimiter=',')
        for row in reader:
            fileName_docID_dict[row[1]] = row[2]

    return fileName_docID_dict

def get_semantic_tag(headerList):
    return Semantic().tagged_list_header(headerList)

def get_regularized_value(value):
    return ValueRegularizer().regularize_value(value)


def get_regularized_value_with_header(hv):
    return ValueRegularizer().regularize_value_with_header(hv)

def ke_bulk(xml_list, OUT_DIR_PATH):
    failed_xml_list = []

    for xml_file in xml_list:
        try:
            doc_id = fileName_docID_dict[xml_file]
            print("XML file", xml_file)
            print("Doc ID", doc_id)
            result = extract_preKB(BASE_DIR_PATH+xml_file, doc_id).toJson()

            # write json
            with open(OUT_DIR_PATH + "preKB_" + xml_file.split(".")[0] + ".json", 'w') as outFile:
                outFile.write(result)
            
            # write excel
            preKB_to_xlsx(result, OUT_DIR_PATH+"preKB_"+xml_file.split(".")[0]+'.xlsx')
        except:
            failed_xml_list.append(xml_file)

    if failed_xml_list:
        with open(OUT_DIR_PATH + "failed_xml_list.txt", 'w') as outFile:
                for xml in failed_xml_list:
                    outFile.write(xml+"\n")


if __name__ == "__main__":

    """
    fileName_docID map
    """
    fileName_docID_dict = get_fileName_docID_dict()
    
    """
    0. 지식 추출 (벌크)
    """
    BASE_DIR_PATH = "/data1/saltlux/CJPoc/table-extraction/final_demo/task1_30/task1_xml_30/"
    OUT_DIR_PATH = "/data1/saltlux/CJPoc/table-extraction/final_demo/task1_30/task1_table_result_30/"
    # BASE_DIR_PATH = "/data1/saltlux/CJPoc/table-extraction/final_demo/task2_216/xml_138/"
    # OUT_DIR_PATH = "/data1/saltlux/CJPoc/table-extraction/final_demo/task2_216/task2_result_138/"
    # BASE_DIR_PATH = "/data1/saltlux/CJPoc/table-extraction/final_demo/task2_216/xml_cj_78/"
    # OUT_DIR_PATH = "/data1/saltlux/CJPoc/table-extraction/final_demo/task2_216/task2_result_78/"
    
    # Base directory
    xml_list = os.listdir(BASE_DIR_PATH)

    # Failed list
    # xml_list = []
    # with open(OUT_DIR_PATH + "failed_xml_list.txt") as failed_file:
    #     for line in failed_file:
    #         xml_list.append(line.rstrip())

    ke_bulk(xml_list, OUT_DIR_PATH)


    """
    지식 추출 - 개별 테스트
    """
    
    # BASE_DIR_PATH = "/data1/saltlux/CJPoc/table-extraction/final_demo/task2_216/xml_cj_78/"
    # xml_file = "2018_S016882271731608X_Serum adiponectinFerritin rati.xml"
    
    # # Table
    # result = PreKB(0, from_structured(BASE_DIR_PATH+xml_file)).toJson()
    # print(result)

    # # Text
    # doc_id = fileName_docID_dict[xml_file]
    # result = extract_preKB(BASE_DIR_PATH+xml_file, doc_id).toJson()
    # print(result)
    # preKB_to_xlsx(result, OUT_DIR_PATH+"preKB_"+xml_file.split(".")[0]+'.xlsx')