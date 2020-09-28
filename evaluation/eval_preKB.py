from openpyxl import Workbook, load_workbook
import os
import collections

def evaluate_table_xlsx(BASE_DIR_PATH, RESULT_PATH):
    
    """
    """

    write_wb = Workbook()
    eval_sheet = write_wb.active
    eval_sheet.title = "Evaluation"
    eval_sheet.append(["paper_name", "groupInfo_precision", "experimentInfo_precision", "precision"])    

    file_list = os.listdir(BASE_DIR_PATH)
    groupInfo_precision_list = []
    groupInfo_true_list = []
    groupInfo_answers_list = []
    experimentInfo_precision_list = []
    experimentInfo_true_list = []
    experimentInfo_answers_list = []
    paper_precision_list = []
    paper_true_list = []
    paper_answers_list = []


    for file_name in sorted(file_list):
        file_path = BASE_DIR_PATH + file_name

        wb = load_workbook(filename = file_path)
        sheet_groupInfo = wb['GroupInfo']
        sheet_experimentInfo = wb['experimentInfo']
        
        # groupInfo
        print(file_name)
        print("groupInfo")
        group_number_of_True = 0
        group_number_of_answers = 0
        try:
            groupInfo_count = collections.Counter([ cell.value for cell in sheet_groupInfo['L'][1:] ])
            print(groupInfo_count)
        
            if 'O' in groupInfo_count.keys() or 'o' in groupInfo_count.keys():
                group_number_of_True += groupInfo_count['O']
                group_number_of_answers += groupInfo_count['O']
            
            if 'X' in groupInfo_count.keys():
                group_number_of_answers += groupInfo_count['X']

            if '△' in groupInfo_count.keys():
                group_number_of_True += groupInfo_count['△']
                group_number_of_answers += groupInfo_count['△']
             
        except: 
            pass

        # experiementInfo
        print("experimentInfo")
        experiment_number_of_True = 0
        experiment_number_of_answers = 0
        try:
            experimentInfo_count = collections.Counter([ cell.value for cell in sheet_experimentInfo['I'][1:] ])
            print(experimentInfo_count)

            if 'O' in experimentInfo_count.keys():
                experiment_number_of_True += experimentInfo_count['O']
                experiment_number_of_answers += experimentInfo_count['O']
            
            if 'X' in experimentInfo_count.keys():
                experiment_number_of_answers += experimentInfo_count['X']

            if '△' in experimentInfo_count.keys():
                experiment_number_of_True += experimentInfo_count['△']
                experiment_number_of_answers += experimentInfo_count['△']
        except:
            pass

        groupInfo_precision = None if group_number_of_answers == 0 else group_number_of_True / group_number_of_answers
        experimentInfo_precision = None if experiment_number_of_answers == 0 else experiment_number_of_True / experiment_number_of_answers
        
        paper_number_of_True = group_number_of_True + experiment_number_of_True
        paper_number_of_ansers = group_number_of_answers + experiment_number_of_answers
        paper_precision = None if paper_number_of_ansers == 0 else paper_number_of_True / paper_number_of_ansers

        # groupInfo
        groupInfo_precision_list.append(groupInfo_precision)
        groupInfo_true_list.append(group_number_of_True)
        groupInfo_answers_list.append(group_number_of_answers)
        # experimentInfo
        experimentInfo_precision_list.append(experimentInfo_precision)
        experimentInfo_true_list.append(experiment_number_of_True)
        experimentInfo_answers_list.append(experiment_number_of_answers)
        # paper
        paper_precision_list.append(paper_precision)
        paper_true_list.append(paper_number_of_True)
        paper_answers_list.append(paper_number_of_ansers)

        # Evaluation 엑셀에 저장
        eval_sheet.append([file_name, groupInfo_precision, experimentInfo_precision, paper_precision])

    # precision == None 인 결과 skip (지식을 추출하지 않은 것)
    # groupInfo
    groupInfo_precision_list = list( filter(lambda x : x is not None, groupInfo_precision_list))
    groupInfo_true_list = list( filter(lambda x : x is not None, groupInfo_true_list))
    groupInfo_answers_list = list( filter(lambda x : x is not None, groupInfo_answers_list))
    # experimentInfo
    experimentInfo_precision_list = list( filter(lambda x : x is not None, experimentInfo_precision_list))
    experimentInfo_true_list = list( filter(lambda x : x is not None, experimentInfo_true_list))
    experimentInfo_answers_list = list( filter(lambda x : x is not None, experimentInfo_answers_list))
    # paper
    paper_precision_list = list( filter(lambda x : x is not None, paper_precision_list))
    paper_true_list = list( filter(lambda x : x is not None, paper_true_list))
    paper_answers_list = list( filter(lambda x : x is not None, paper_answers_list))

    
    # avarage of precisions    
    # groupInfo_precision_total = sum(groupInfo_precision_list)/len(groupInfo_precision_list)
    # experimentInfo_precision_total = sum(experimentInfo_precision_list)/len(experimentInfo_precision_list)
    #paper_precision_total = sum(paper_precision_list)/len(paper_precision_list)

    # total collect / total answers
    groupInfo_precision_total = sum(groupInfo_true_list)/sum(groupInfo_answers_list)
    experimentInfo_precision_total = sum(experimentInfo_true_list)/sum(experimentInfo_answers_list)
    paper_precision_total = sum(paper_true_list)/sum(paper_answers_list)

    eval_sheet.append(["Total", groupInfo_precision_total, experimentInfo_precision_total, paper_precision_total])    
    
    write_wb.save(RESULT_PATH)


def evaluate_text_xlsx(BASE_DIR_PATH, RESULT_PATH, isTriangleCollect = True):
    
    """
    """

    write_wb = Workbook()
    eval_sheet = write_wb.active
    eval_sheet.title = "Evaluation"
    eval_sheet.append(["paper_name", "precision"]) 

    file_list = os.listdir(BASE_DIR_PATH)
    tgInfo_precision_list = []
    tgInfo_true_list = []
    tgInfo_answers_list = []


    for file_name in sorted(file_list):
        file_path = BASE_DIR_PATH + file_name

        wb = load_workbook(filename = file_path)
        sheet_treatmentGeneInfo = wb['treatmentGeneInfo']
        
        # groupInfo
        print(file_name)
        print("treatmentGeneInfo")
        tgInfo_number_of_True = 0
        tgInfo_number_of_answers = 0
        try:
            tgInfo_count = collections.Counter([ cell.value for cell in sheet_treatmentGeneInfo['N'][1:] ])
            print(tgInfo_count)
        
            if 'O' in tgInfo_count.keys():
                tgInfo_number_of_True += tgInfo_count['O']
                tgInfo_number_of_answers += tgInfo_count['O']
            
            if 'X' in tgInfo_count.keys():
                tgInfo_number_of_answers += tgInfo_count['X']

            if '△' in tgInfo_count.keys() and isTriangleCollect:
                tgInfo_number_of_True += tgInfo_count['△']
                tgInfo_number_of_answers += tgInfo_count['△']
            elif '△' in tgInfo_count.keys() and isTriangleCollect is not True:
                tgInfo_number_of_answers += tgInfo_count['△']
             
        except: 
            pass
        
        tgInfo_precision = None if tgInfo_number_of_answers == 0 else tgInfo_number_of_True / tgInfo_number_of_answers
        
        tgInfo_true_list.append(tgInfo_number_of_True)
        tgInfo_answers_list.append(tgInfo_number_of_answers)
        tgInfo_precision_list.append(tgInfo_precision)

        # Evaluation 엑셀에 저장
        eval_sheet.append([file_name, tgInfo_precision])

    # precision == None 인 결과 skip (지식을 추출하지 않은 것)
    tgInfo_precision_list = list( filter(lambda x : x is not None, tgInfo_precision_list))
    tgInfo_true_list = list( filter(lambda x : x is not None, tgInfo_true_list))
    tgInfo_answers_list = list( filter(lambda x : x is not None, tgInfo_answers_list))
    
    # avarage of precisions
    #tgInfo_precision_total = sum(tgInfo_precision_list)/len(tgInfo_precision_list)

    # total collect / total answers
    tgInfo_precision_total = sum(tgInfo_true_list)/sum(tgInfo_answers_list)
    
    eval_sheet.append(["Total", tgInfo_precision_total])    
    
    write_wb.save(RESULT_PATH)

if __name__ == "__main__":

    TABLE_BASE_DIR_PATH = "/data1/saltlux/CJPoc/table-extraction/evaluation/data/표 지식 추출_preKB 평가 결과_200928(216개)/"
    TEXT_BASE_DIR_PATH = "/data1/saltlux/CJPoc/table-extraction/evaluation/data/비정형 지식 추출_preKB 평가 결과_200928(216개)/비정형 지식 추출_preKB 평가 결과_200928(174개)/"
    TABLE_RESULT_PATH = "/data1/saltlux/CJPoc/table-extraction/evaluation/result/evaluation_table.xlsx"
    TEXT_RESULT_PATH = "/data1/saltlux/CJPoc/table-extraction/evaluation/result/evaluation_text_triangle_true.xlsx"
    TEXT_RESULT_PATH_2 = "/data1/saltlux/CJPoc/table-extraction/evaluation/result/evaluation_text_triangle_false.xlsx"


    evaluate_table_xlsx(TABLE_BASE_DIR_PATH, TABLE_RESULT_PATH)
    #evaluate_text_xlsx(TEXT_BASE_DIR_PATH, TEXT_RESULT_PATH, True)
    #evaluate_text_xlsx(TEXT_BASE_DIR_PATH, TEXT_RESULT_PATH_2, False)